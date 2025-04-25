import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.title("Excel File Reader with Month and Year Filter")




WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"

if 'processed_df' not in st.session_state:
    st.session_state.processed_df = None
if 'total_count_df' not in st.session_state:
    st.session_state.total_count_df = None
if 'selected_file_name' not in st.session_state:
    st.session_state.selected_file_name = None
if 'sheduledf' not in st.session_state:
    st.session_state.sheduledf = None
if 'shedule' not in st.session_state:
    st.session_state.shedule = None
if 'df_selected' not in st.session_state:
    st.session_state.df_selected = None

def GetAccesstoken():
    auth_url = "https://iam.cloud.ibm.com/identity/token"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Accept": "application/json"
    }
    data = {
        "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
        "apikey": API_KEY
    }
    response = requests.post(auth_url, headers=headers, data=data)
    if response.status_code != 200:
        st.error(f"Failed to get access token: {response.text}")
        return None
    return response.json()['access_token']

def generatePrompt(json_datas):
    body = {
        "input": f"""
        read this table properly and i need total count of each acitvity as json example provided
        {json_datas}

        Example json needed:
        [{{
            "Activity Name":"name",
            "Total":"Count"
        }}]
Return only the JSON object, no code, no explanation, just the formatted JSON, and count properly please.
        """, 
        "parameters": {
            "decoding_method": "greedy",
            "max_new_tokens": 8100,
            "min_new_tokens": 0,
            "stop_sequences": [";"],
            "repetition_penalty": 1.05,
            "temperature": 0.5
        },
        "model_id": MODEL_ID,
        "project_id": PROJECT_ID
    }
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GetAccesstoken()}"
    }
    if not headers["Authorization"]:
        return "Error: No valid access token."
    response = requests.post(WATSONX_API_URL, headers=headers, json=body)
    if response.status_code != 200:
        st.error(f"Failed to generate prompt: {response.text}")
        return "Error generating prompt"
    return response.json()['results'][0]['generated_text'].strip()



# def to_excel(df):
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         df.to_excel(writer, index=True, sheet_name='Activity Counts')
#     return output.getvalue()

def to_excel(df, year):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write DataFrame starting from row 2 (Excel row 2, 1-based)
        df.to_excel(writer, index=True, sheet_name='Activity Counts', startrow=1)
        
        # Get the openpyxl workbook and worksheet
        workbook = writer.book
        worksheet = workbook['Activity Counts']
        
        # Define the title
        title = f"Activity Counts Report:({year})"
        
        # Calculate the number of columns (index + DataFrame columns)
        total_columns = len(df.columns) + 1  # +1 for index column
        
        # Merge cells in the first row for the title
        start_cell = 'A1'
        end_cell = f'{get_column_letter(total_columns)}1'
        worksheet.merge_cells(f'{start_cell}:{end_cell}')
        
        # Set the title text
        title_cell = worksheet['A1']
        title_cell.value = title
        
        # Apply styling: yellow background, bold font, centered
        title_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        title_cell.font = Font(bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
    return output.getvalue()

def get_cos_files():
    try:
        response = st.session_state.cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        if not files:
            st.warning("No .xlsx files found in the bucket 'projectreport'. Please ensure Excel files are uploaded.")
        return files
    except Exception as e:
        st.error(f"Error fetching COS files: {e}")
        return []

def getTotal(ai_data):
    share = []
    for i in ai_data:
        share.append(i['Total'])
    return share

# Function to process the Excel file
def process_file(file_stream):
    workbook = openpyxl.load_workbook(file_stream)
    
    if "M7 T5" in workbook.sheetnames:
        df = pd.read_excel(file_stream, sheet_name="M7 T5", header=1)
        target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Start ()", "Finish ()"]
        existing_columns = [col for col in target_columns if col in df.columns]

        activity_col_idx = df.columns.get_loc("Activity Name") + 1

        non_bold_indices = []
        for i, row in enumerate(workbook["M7 T5"].iter_rows(min_row=3, max_row=workbook["M7 T5"].max_row), start=0):
            cell = row[activity_col_idx - 1]
            if not (cell.font and cell.font.bold):
                non_bold_indices.append(i)

        df_non_bold = df.iloc[non_bold_indices]
        df_selected = df_non_bold[existing_columns]

        # Convert 'Finish ()' column to datetime
        df_selected['Finish ()'] = pd.to_datetime(df_selected['Finish ()'], errors='coerce')

        # Extract Month Name and Year
        df_selected['Finish Month'] = df_selected['Finish ()'].dt.strftime('%b')
        df_selected['Finish Year'] = df_selected['Finish ()'].dt.year

        return df_selected
    
    elif "TOWER 4 FINISHING." in workbook.sheetnames:
        sheet_name = "TOWER 4 FINISHING."
        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=1)
        
        # Assign column names
        df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
                      'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                      'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
        
        # Select desired columns
        target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Start", "Finish"]
        df = df[target_columns]
        
        # Define the column index for 'Activity Name' (0-based index, column F = index 5)
        activity_col_idx = 5
        
        # Get row indices where Activity Name is not bold
        non_bold_rows = [
            row_idx for row_idx, row in enumerate(workbook[sheet_name].iter_rows(min_row=2, max_col=16), start=1)
            if row[activity_col_idx].font is not None and not row[activity_col_idx].font.b
        ]
        
        # Adjust row indices to match pandas DataFrame indexing
        # header=1 means Excel row 3 = DataFrame index 0, so subtract 2
        non_bold_rows = [idx - 2 for idx in non_bold_rows]
        
        # Filter valid indices to prevent out-of-bounds error
        max_index = len(df) - 1
        non_bold_rows = [idx for idx in non_bold_rows if 0 <= idx <= max_index]
        
        # Filter the DataFrame to include only rows where Activity Name is not bold
        if non_bold_rows:
            df_non_bold = df.iloc[non_bold_rows]
        else:
            df_non_bold = pd.DataFrame(columns=df.columns)
        
        # Convert 'Finish' column to datetime
        df_non_bold['Finish'] = pd.to_datetime(df_non_bold['Finish'], errors='coerce')
        
        # Extract Month Name and Year
        df_non_bold['Finish Month'] = df_non_bold['Finish'].dt.strftime('%b')
        df_non_bold['Finish Year'] = df_non_bold['Finish'].dt.year
        
        return df_non_bold
    
    else:
        return None

# Streamlit App
st.title("Excel File Activity Processor")

# Initialize session state
if 'df_selected' not in st.session_state:
    st.session_state.df_selected = None
if 'selected_file' not in st.session_state:
    st.session_state.selected_file = None

# Get files from COS
files = get_cos_files()
st.sidebar.header("Select a File")
selected_file = st.sidebar.selectbox("Choose a file from IBM COS:", files, key="file_selector")

# Process file only if a new file is selected
if selected_file and selected_file != st.session_state.selected_file:
    st.session_state.selected_file = selected_file
    response = st.session_state.cos_client.get_object(Bucket="projectreportnew", Key=selected_file)
    st.session_state.df_selected = process_file(io.BytesIO(response['Body'].read()))

# Work with the stored DataFrame
df_selected = st.session_state.df_selected

if df_selected is None:
    st.write("No valid data found in the selected file. Ensure the file contains 'M7 T5' or 'TOWER 4 FINISHING.' sheet.")
else:
    st.write("Processed Data:")
    st.write(df_selected)

    # Get unique years and months for filters
    unique_years = sorted(df_selected['Finish Year'].dropna().unique().astype(int))
    unique_months = sorted(df_selected['Finish Month'].dropna().unique())

    # Sidebar filters
    st.sidebar.header("Filters")
    selected_year = st.sidebar.selectbox("Select Year", unique_years, key="year_filter")
    selected_months = st.sidebar.multiselect("Select Months", unique_months, default=unique_months, key="month_filter")

    # Apply filters directly on the stored DataFrame
    filtered_data = df_selected[
        (df_selected['Finish Year'] == selected_year) & 
        (df_selected['Finish Month'].isin(selected_months))
    ]

    st.subheader(f"Filtered Data for {', '.join(selected_months)} {selected_year}")
    st.write(filtered_data.head(20))

    if st.button("Display Activity Count by Month"):
        # Get only necessary columns for activity and month
        activity_month_data = filtered_data[['Activity Name', 'Finish Month']]

        # Group and count activities by month
        count_table = (
            activity_month_data
            .groupby(['Activity Name', 'Finish Month'])
            .size()
            .reset_index(name='Count')
            .pivot(index='Activity Name', columns='Finish Month', values='Count')
            .fillna(0)
            .astype(int)
        )

        # Sort months in calendar order
        month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        count_table = count_table.reindex(columns=sorted(count_table.columns, key=lambda x: month_order.index(x)))

        # Get AI-generated data (assuming generatePrompt and getTotal are defined elsewhere)
        test = generatePrompt(count_table)
        ai_data = json.loads(test)
        count_table['Total'] = getTotal(ai_data)

        st.write("Activity Count by Month:")
        st.write(count_table)
        st.session_state.sheduledf = count_table
        st.session_state.shedule = to_excel(count_table, selected_year)
        st.download_button(
            label="📥 Download Excel Report",
            data=st.session_state.shedule,
            file_name=f"Shedule_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )  
        # st.write("AI-Generated Data:")
        # st.write(ai_data)