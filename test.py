# import streamlit as st
# import pandas as pd
# from openpyxl import load_workbook
# import requests
# import json

# WATSONX_API_URL = "https://us-south.ml.cloud.ibm.com/ml/v1/text/generation?version=2023-05-29"
# MODEL_ID = "meta-llama/llama-3-2-90b-vision-instruct"
# PROJECT_ID = "4152f31e-6a49-40aa-9b62-0ecf629aae42"
# API_KEY = "KEmIMzkw273qBcek8IdF-aShRUvFwH7K4psARTqOvNjI"


# def GetAccesstoken():
#     auth_url = "https://iam.cloud.ibm.com/identity/token"
#     headers = {
#         "Content-Type": "application/x-www-form-urlencoded",
#         "Accept": "application/json"
#     }
#     data = {
#         "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
#         "apikey": API_KEY  # Define API_KEY
#     }
#     response = requests.post(auth_url, headers=headers, data=data)
#     if response.status_code != 200:
#         st.error(f"Failed to get access token: {response.text}")
#         return None
#     return response.json()['access_token']

# # Generate prompt for Watson API
# def generatePrompt(json_datas):
#     body = {
#         "input": f"""
#         read this table properly and i need total count of each acitvity as json example provided
#         {json_datas}

#         Example json needed:
#         [{{
#             'Activity Name':'name',
#             'Total':'Count'
#         }}]
# Return only the JSON object, no code, no explanation, just the formatted JSON, and count properly please.
#         """, 
#         "parameters": {
#             "decoding_method": "greedy",
#             "max_new_tokens": 8100,
#             "min_new_tokens": 0,
#             "stop_sequences": [";"],
#             "repetition_penalty": 1.05,
#             "temperature": 0.5
#         },
#         "model_id": MODEL_ID,  # Define MODEL_ID
#         "project_id": PROJECT_ID  # Define PROJECT_ID
#     }
#     headers = {
#         "Accept": "application/json",
#         "Content-Type": "application/json",
#         "Authorization": f"Bearer {GetAccesstoken()}"
#     }
#     if not headers["Authorization"]:
#         return "Error: No valid access token."
#     response = requests.post(WATSONX_API_URL, headers=headers, json=body)  # Define WATSONX_API_URL
#     if response.status_code != 200:
#         st.error(f"Failed to generate prompt: {response.text}")
#         return "Error generating prompt"
#     return response.json()['results'][0]['generated_text'].strip()


# # Title
# st.title("Excel Viewer: Non-Bold Filter, Selected Columns & Month/Year Extraction")

# # Sidebar for file uploader
# uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "xls"])


# def getTotal(datas):
#     share = []
#     for i in datas:
#         share.append(i['Total'])
#     return share

# if uploaded_file is not None:
#     try:
       
#         if 'df_selected' not in st.session_state:
          
#             wb = load_workbook(uploaded_file, data_only=True)
#             ws = wb["TOWER 4 FINISHING."]

            
#             df = pd.read_excel(uploaded_file, sheet_name="TOWER 4 FINISHING.")
#             df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
#                       'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
#                       'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
#             df = df[['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']]
#             st.write(df)


#     except Exception as e:
#         st.error(f"Error processing the Excel file: {e}")
import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import io
from datetime import datetime

# Function to extract year and month from a date
def extract_year_month(date):
    try:
        if pd.isna(date):
            return None, None
        if isinstance(date, (int, float)):  # Excel serial date
            date = pd.to_datetime(date, origin='1899-12-30', unit='D')
        elif isinstance(date, str):
            date = pd.to_datetime(date)
        return date.year, date.month
    except (ValueError, TypeError):
        return None, None

# File uploader for Excel file
uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"], key="file_uploader")

# Initialize session state
if 'df' not in st.session_state:
    st.session_state['df'] = None
if 'non_bold_df' not in st.session_state:
    st.session_state['non_bold_df'] = None
if 'filtered_df' not in st.session_state:
    st.session_state['filtered_df'] = None
if 'file_processed' not in st.session_state:
    st.session_state['file_processed'] = False

if uploaded_file is not None and not st.session_state['file_processed']:
    try:
        # Load the workbook
        workbook = load_workbook(uploaded_file, data_only=True)

        # Check if the sheet "TOWER 4 FINISHING." exists
        if "TOWER 4 FINISHING." in workbook.sheetnames:
            sheet_name = "TOWER 4 FINISHING."
            sheet = workbook[sheet_name]

            # Read the Excel sheet into a DataFrame
            df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=1, engine='openpyxl')
            
            # Assign column names
            df.columns = ['Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name', 
                          'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish', 
                          'Actual Start', 'Actual Finish', '%Complete', 'Start', 'Finish', 'Delay Reasons']
            
            # Select desired columns
            df = df[['Module', 'Floor', 'Flat', 'Activity ID', 'Activity Name', 'Start', 'Finish']]
            
            # Debug: Display DataFrame size
            st.write(f"DataFrame shape: {df.shape}")
            
            # Define the column index for 'Activity Name' (0-based index, column F = index 5)
            activity_col_idx = 5
            
            # Get row indices where Activity Name is not bold
            non_bold_rows = [
                row_idx for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=16), start=1)
                if row[activity_col_idx].font is not None and not row[activity_col_idx].font.b
            ]
            
            # Debug: Display non-bold rows before adjustment
            st.write(f"Non-bold rows (Excel indices): {non_bold_rows}")
            
            # Adjust row indices to match pandas DataFrame indexing
            # header=1 means Excel row 3 = DataFrame index 0, so subtract 2
            non_bold_rows = [idx - 2 for idx in non_bold_rows]
            
            # Debug: Display non-bold rows after adjustment
            st.write(f"Non-bold rows (DataFrame indices): {non_bold_rows}")
            
            # Filter valid indices to prevent out-of-bounds error
            max_index = len(df) - 1
            non_bold_rows = [idx for idx in non_bold_rows if 0 <= idx <= max_index]
            
            # Debug: Display filtered non-bold rows
            st.write(f"Filtered non-bold rows (within bounds): {non_bold_rows}")
            
            # Filter the DataFrame to include only rows where Activity Name is not bold
            if non_bold_rows:
                non_bold_df = df.iloc[non_bold_rows]
            else:
                non_bold_df = pd.DataFrame(columns=df.columns)
            
            # Store in session state
            st.session_state['df'] = df
            st.session_state['non_bold_df'] = non_bold_df
            st.session_state['file_processed'] = True
            
        else:
            st.error("Sheet 'TOWER 4 FINISHING.' not found in the workbook.")
            st.session_state['file_processed'] = False
    except Exception as e:
        st.error(f"Error processing the Excel file: {str(e)}")
        st.session_state['file_processed'] = False

# Year and Month Filters
if st.session_state['non_bold_df'] is not None and not st.session_state['non_bold_df'].empty:
    # Extract unique years from Start and Finish
    years = set()
    for col in ['Start', 'Finish']:
        for date in st.session_state['non_bold_df'][col]:
            year, _ = extract_year_month(date)
            if year is not None:
                years.add(year)
    years = sorted(list(years))
    
    # Month names for selection
    months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    
    # Create filter inputs
    st.subheader("Filter by Year and Month")
    col1, col2 = st.columns(2)
    with col1:
        selected_year = st.selectbox("Select Year", options=years, key="year_filter")
    with col2:
        selected_month = st.selectbox("Select Month", options=months, key="month_filter")
    
    # Convert selected month to month number (1-12)
    month_num = months.index(selected_month) + 1
    
    # Filter non_bold_df based on year and month
    filtered_df = st.session_state['non_bold_df'].copy()
    mask = pd.Series(False, index=filtered_df.index)
    
    for col in ['Start', 'Finish']:
        for idx, date in filtered_df[col].items():
            year, month = extract_year_month(date)
            if year == selected_year and month == month_num:
                mask[idx] = True
    
    filtered_df = filtered_df[mask]
    
    # Store filtered DataFrame in session state
    st.session_state['filtered_df'] = filtered_df
    
    # Display filtered DataFrame
    st.write(f"Filtered DataFrame (Year: {selected_year}, Month: {selected_month}):")
    st.write(st.session_state['filtered_df'])
    
    # Download button for filtered DataFrame
    if not st.session_state['filtered_df'].empty:
        csv = st.session_state['filtered_df'].to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Download filtered data as CSV",
            data=csv,
            file_name=f"non_bold_activities_{selected_year}_{selected_month}.csv",
            mime="text/csv"
        )
    else:
        st.warning("No data matches the selected year and month.")
else:
    if uploaded_file is not None and st.session_state['file_processed']:
        st.warning("No non-bold Activity Names found in the data.")
    else:
        st.info("Please upload an Excel file to begin.")

